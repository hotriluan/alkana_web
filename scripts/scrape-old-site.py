"""
scrape-old-site.py — Alkana old website scraper.

Crawls the old alkana.vn website and exports product data to:
  raw-alkana-products.xlsx

Usage:
    python scrape-old-site.py [--url https://www.alkana.vn] [--output raw-alkana-products.xlsx]

Requirements:
    pip install requests beautifulsoup4 openpyxl lxml

IMPORTANT: Run at <= 1 request/second to avoid triggering DDoS protection.
Old site check first: request index page to confirm site is up before batch scraping.
"""

import argparse
import os
import re
import time
import logging
from typing import Optional
from urllib.parse import urljoin, urlparse

import requests
from bs4 import BeautifulSoup
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

# ── Config ────────────────────────────────────────────────────────────────────

BASE_URL       = "https://www.alkana.vn"
PRODUCT_PATHS  = ["/san-pham/", "/product/", "/products/"]
REQUEST_DELAY  = 1.0   # seconds between requests
REQUEST_TIMEOUT = 15   # seconds
MAX_PRODUCTS   = 9999  # safety cap

HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (compatible; AlkanaWebMigration/1.0; "
        "+https://github.com/hotriluan/alkana_web)"
    ),
    "Accept-Language": "vi-VN,vi;q=0.9,en;q=0.8",
}

EXCEL_COLUMNS = [
    # Scraped columns
    ("A", "old_url",        "Old URL (scraped)"),
    ("B", "title",          "Product title (scraped)"),
    ("C", "description",    "Description (scraped)"),
    ("D", "sku",            "SKU — fill in or confirm"),
    ("E", "image_url",      "Primary image URL (scraped)"),
    ("F", "tds_url",        "TDS PDF URL (scraped)"),
    ("G", "msds_url",       "MSDS PDF URL (scraped)"),
    ("H", "old_category",   "Old category (scraped)"),
    # Columns for Alkana team to fill / confirm
    ("I",  "new_category_slug", "New category slug (fill in)"),
    ("J",  "surface_types",     "Surface types — comma sep slugs (fill in)"),
    ("K",  "paint_system",      "Paint system slug (fill in)"),
    ("L",  "gloss_level",       "Gloss level slug (fill in)"),
    ("M",  "coverage",          "Coverage e.g. 8–10 m²/L (fill in)"),
    ("N",  "mix_ratio",         "Mix ratio e.g. 4:1 (fill in)"),
    ("O",  "thinner",           "Thinner / dilution (fill in)"),
    ("P",  "dry_touch",         "Dry to touch time (fill in)"),
    ("Q",  "dry_hard",          "Dry hard time (fill in)"),
    ("R",  "dry_recoat",        "Recoat interval (fill in)"),
    ("S",  "new_url_slug",      "New URL slug for 301 map (fill in)"),
    ("T",  "notes",             "Notes / flags for import team"),
]

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    datefmt="%H:%M:%S",
)
log = logging.getLogger("alkana-scraper")


# ── HTTP helper ───────────────────────────────────────────────────────────────

session = requests.Session()
session.headers.update(HEADERS)


def fetch(url: str) -> Optional[BeautifulSoup]:
    """Fetch URL and return BeautifulSoup, or None on error."""
    try:
        time.sleep(REQUEST_DELAY)
        r = session.get(url, timeout=REQUEST_TIMEOUT, allow_redirects=True)
        r.raise_for_status()
        return BeautifulSoup(r.text, "lxml")
    except requests.RequestException as e:
        log.warning("Failed to fetch %s: %s", url, e)
        return None


# ── Sitemap / URL discovery ───────────────────────────────────────────────────

def discover_product_urls(base: str) -> list[str]:
    """
    Try sitemap.xml first; fall back to crawling the products archive page.
    Returns a deduplicated list of absolute product URLs.
    """
    urls: set[str] = set()

    # Strategy 1 — XML sitemap
    for sitemap_path in ("/sitemap.xml", "/sitemap_index.xml", "/sitemap-products.xml"):
        sitemap_url = base.rstrip("/") + sitemap_path
        log.info("Trying sitemap: %s", sitemap_url)
        soup = fetch(sitemap_url)
        if not soup:
            continue
        # sitemap uses <loc> tags
        locs = [tag.get_text(strip=True) for tag in soup.find_all("loc")]
        product_locs = [u for u in locs if any(p.strip("/") in u for p in PRODUCT_PATHS)]
        if product_locs:
            log.info("Found %d product URLs in sitemap.", len(product_locs))
            urls.update(product_locs)
            break

    # Strategy 2 — Paginated archive crawl
    if not urls:
        for archive_path in PRODUCT_PATHS:
            archive_url = base.rstrip("/") + archive_path
            log.info("Crawling archive: %s", archive_url)
            page = 1
            while True:
                page_url = f"{archive_url}page/{page}/" if page > 1 else archive_url
                soup = fetch(page_url)
                if not soup:
                    break
                links = soup.select("a[href]")
                found = 0
                for a in links:
                    href = urljoin(base, a["href"])
                    # Include only URLs that look like product detail pages
                    if any(href.startswith(base + p) for p in PRODUCT_PATHS):
                        parsed = urlparse(href)
                        # Must have a path depth of >= 3 (archive/category/product)
                        parts = [p for p in parsed.path.strip("/").split("/") if p]
                        if len(parts) >= 2:
                            urls.add(href)
                            found += 1
                if not found or page > 100:
                    break
                page += 1
            if urls:
                break

    return sorted(urls)


# ── Product parser ────────────────────────────────────────────────────────────

def _text(soup: BeautifulSoup, *selectors: str) -> str:
    """Try multiple CSS selectors, return first non-empty text."""
    for sel in selectors:
        el = soup.select_one(sel)
        if el:
            return el.get_text(separator=" ", strip=True)
    return ""


def _attr(soup: BeautifulSoup, selector: str, attr: str) -> str:
    el = soup.select_one(selector)
    return (el.get(attr) or "") if el else ""


def _pdf_url(soup: BeautifulSoup, keyword: str, base: str) -> str:
    """Find first PDF link whose href or anchor text contains keyword."""
    for a in soup.find_all("a", href=True):
        href = a["href"]
        text = a.get_text(strip=True).upper()
        if ".pdf" in href.lower() and keyword.upper() in (href.upper() + text):
            return urljoin(base, href)
    return ""


def scrape_product(url: str) -> dict:
    """Scrape a single product page. Returns a dict of raw scraped data."""
    soup = fetch(url)
    if not soup:
        return {"old_url": url, "title": "ERROR: could not fetch", "notes": "fetch_failed"}

    # Title — try common selectors
    title = _text(soup,
        "h1.product-title", "h1.product_title", "h1.entry-title", "h1",
    )

    # Description
    description = _text(soup,
        ".product-description", ".woocommerce-product-details__short-description",
        ".entry-content", ".product-content", ".description",
    )
    # Truncate very long descriptions
    description = description[:1000] if description else ""

    # Primary image
    image_url = (
        _attr(soup, ".wp-post-image",          "src")
        or _attr(soup, ".product-image img",   "src")
        or _attr(soup, ".product__image img",  "src")
        or _attr(soup, ".entry-thumb img",     "src")
        or _attr(soup, "article img",          "src")
    )
    if image_url:
        image_url = urljoin(url, image_url)

    # PDFs
    tds_url  = _pdf_url(soup, "TDS",  url) or _pdf_url(soup, "technical", url)
    msds_url = _pdf_url(soup, "MSDS", url) or _pdf_url(soup, "SDS",  url) or _pdf_url(soup, "safety", url)

    # Old category
    category = _text(soup,
        ".product-category a", ".product-cat a",
        "span.posted_in a", ".woocommerce-breadcrumb a:last-child",
    )

    # Auto-generate new URL slug from title
    slug = re.sub(r"[^a-z0-9]+", "-", title.lower().strip()).strip("-")

    return {
        "old_url":      url,
        "title":        title,
        "description":  description,
        "sku":          "",
        "image_url":    image_url,
        "tds_url":      tds_url,
        "msds_url":     msds_url,
        "old_category": category,
        # Columns for Alkana team
        "new_category_slug": "",
        "surface_types":     "",
        "paint_system":      "",
        "gloss_level":       "",
        "coverage":          "",
        "mix_ratio":         "",
        "thinner":           "",
        "dry_touch":         "",
        "dry_hard":          "",
        "dry_recoat":        "",
        "new_url_slug":      slug,
        "notes":             "",
    }


# ── Excel export ──────────────────────────────────────────────────────────────

def export_to_excel(products: list[dict], output_path: str) -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Products"

    # Header row
    header_fill  = PatternFill("solid", fgColor="1A3C6B")
    scraped_fill = PatternFill("solid", fgColor="E8F4FD")
    manual_fill  = PatternFill("solid", fgColor="FFF8E1")

    for col_letter, field_key, label in EXCEL_COLUMNS:
        cell = ws[f"{col_letter}1"]
        cell.value     = label
        cell.font      = Font(bold=True, color="FFFFFF")
        cell.fill      = header_fill
        cell.alignment = Alignment(wrap_text=True)

    # Data rows
    for i, product in enumerate(products, start=2):
        for col_letter, field_key, _ in EXCEL_COLUMNS:
            cell = ws[f"{col_letter}{i}"]
            cell.value = product.get(field_key, "")
            # Highlight manually-fillable columns
            if col_letter >= "I":
                cell.fill = manual_fill
            else:
                cell.fill = scraped_fill

    # Column widths
    col_widths = {"A": 45, "B": 35, "C": 60, "D": 15, "E": 45,
                  "F": 45, "G": 45, "H": 25, "S": 30, "T": 40}
    for col_letter in "ABCDEFGHIJKLMNOPQRST":
        ws.column_dimensions[col_letter].width = col_widths.get(col_letter, 20)

    ws.freeze_panes = "A2"

    # Legend sheet
    legend = wb.create_sheet("Legend")
    legend["A1"] = "Column"
    legend["B1"] = "Field Key"
    legend["C1"] = "Description"
    legend["D1"] = "Source"
    for row_i, (col_letter, field_key, label) in enumerate(EXCEL_COLUMNS, start=2):
        legend[f"A{row_i}"] = col_letter
        legend[f"B{row_i}"] = field_key
        legend[f"C{row_i}"] = label
        legend[f"D{row_i}"] = "SCRAPED" if col_letter < "I" else "ALKANA TEAM — fill in"

    wb.save(output_path)
    log.info("✅ Saved %d products to %s", len(products), output_path)


# ── Main ──────────────────────────────────────────────────────────────────────

def main() -> None:
    parser = argparse.ArgumentParser(description="Alkana old site product scraper")
    parser.add_argument("--url",    default=BASE_URL,                   help="Base URL of old site")
    parser.add_argument("--output", default="raw-alkana-products.xlsx", help="Output Excel filename")
    parser.add_argument("--limit",  type=int, default=MAX_PRODUCTS,     help="Max products to scrape")
    parser.add_argument("--dry-run", action="store_true",               help="Discover URLs only, no scraping")
    args = parser.parse_args()

    log.info("Alkana product scraper starting — target: %s", args.url)

    # Verify site is reachable
    try:
        r = session.get(args.url, timeout=REQUEST_TIMEOUT)
        r.raise_for_status()
        log.info("Site reachable — HTTP %d", r.status_code)
    except requests.RequestException as e:
        log.error("Cannot reach %s: %s", args.url, e)
        log.error("Tip: ask Alkana team for manual export from old CMS admin.")
        return

    # Discover product URLs
    urls = discover_product_urls(args.url)
    log.info("Discovered %d product URLs.", len(urls))

    if args.dry_run:
        print("\n".join(urls[:50]))
        log.info("Dry run — exiting before scrape.")
        return

    urls = urls[: args.limit]

    # Scrape each product
    products = []
    for i, url in enumerate(urls, start=1):
        log.info("[%d/%d] Scraping: %s", i, len(urls), url)
        data = scrape_product(url)
        products.append(data)

    export_to_excel(products, args.output)
    log.info("Done. Review %s and send to Alkana team for cleansing.", args.output)


if __name__ == "__main__":
    main()
